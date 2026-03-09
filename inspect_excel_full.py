import openpyxl
import os

file_path = r"G:\マイドライブ\Python_code\eBayリサーチ部隊\リサーチシート完全版 のコピー.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet_name = "3月" if "3月" in wb.sheetnames else "3月 のコピー"
    ws = wb[sheet_name]
    print(f"Investigating sheet: {sheet_name}")

    target_row = 2
    cols = ["H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "AG", "AH", "AI"]
    
    print(f"\n--- Row {target_row} Formulas ---")
    for c in cols:
        cell = f"{c}{target_row}"
        val = ws[cell].value
        print(f"{c}: {val}")

    # AI2 (Exchange Rate)
    print(f"\nAI2 (Exchange Rate): {ws['AI2'].value}")

    wb.close()
except Exception as e:
    print(f"Error: {e}")
