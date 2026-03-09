import openpyxl
import os

file_path = r"G:\マイドライブ\Python_code\eBayリサーチ部隊\リサーチシート完全版 のコピー.xlsx"

try:
    wb_f = openpyxl.load_workbook(file_path, data_only=False)
    wb_v = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet_name = "3月" if "3月" in wb_f.sheetnames else "3月 のコピー"
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]
    
    cell_f = ws_f['AG2']
    cell_v = ws_v['AG2']
    
    print(f"--- AG2 Detail Analysis ---")
    print(f"Value (data_only=True): {cell_v.value}")
    print(f"Type (data_only=True): {type(cell_v.value)}")
    
    if hasattr(cell_f.value, 'text'):
        print(f"Array Formula Text: {cell_f.value.text}")
        print(f"Array Formula Ref: {cell_f.value.ref}")
    else:
        print(f"Formula: {cell_f.value}")

    # ついでに近傍のセル（カテゴリー名が入っていそうな列、例えばB列やその他）を確認
    print(f"\n--- Row 2 Context ---")
    for col in ["A", "B", "C", "D", "E", "F", "AE", "AF"]:
        val = ws_v[f"{col}2"].value
        print(f"{col}2: {val}")

    wb_f.close()
    wb_v.close()
except Exception as e:
    print(f"Error: {e}")
