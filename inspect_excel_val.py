import openpyxl
import os

file_path = r"G:\マイドライブ\Python_code\eBayリサーチ部隊\リサーチシート完全版 のコピー.xlsx"

try:
    # 数式用と値用の2つを開く
    wb_f = openpyxl.load_workbook(file_path, data_only=False)
    wb_v = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet_name = "3月" if "3月" in wb_f.sheetnames else "3月 のコピー"
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]
    
    print(f"Investigating sheet: {sheet_name}")

    target_row = 2
    cols = ["G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "X", "AG", "AH", "AI"]
    
    print(f"\n--- Row {target_row} Analysis ---")
    print(f"{'Col':<4} | {'Value':<12} | {'Formula'}")
    print("-" * 50)
    for c in cols:
        cell = f"{c}{target_row}"
        val = ws_v[cell].value
        formula = ws_f[cell].value
        print(f"{c:<4} | {str(val):<12} | {formula}")

    wb_f.close()
    wb_v.close()
except Exception as e:
    print(f"Error: {e}")
