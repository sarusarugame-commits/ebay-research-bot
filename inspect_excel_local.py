import openpyxl
import os

file_path = r"G:\マイドライブ\Python_code\eBayリサーチ部隊\リサーチシート完全版 のコピー.xlsx"

if not os.path.exists(file_path):
    print(f"Error: {file_path} not found")
    exit(1)

try:
    # 読み取り専用で開く（ロック回避のため）が、data_only=False で数式を取得
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    # ユーザー指定は「三月のシート」とのこと。
    sheet_name = "3月" if "3月" in wb.sheetnames else "3月 のコピー"
    ws = wb[sheet_name]
    print(f"Investigating sheet: {sheet_name}")

    # 2行目から数件、数式を確認
    for target_row in [2, 3, 10]:
        columns = {
            "H": "US価格",
            "I": "US内部価格",
            "J": "UK価格",
            "K": "スプシK",
            "L": "スプシL",
            "M": "US利益率",
            "P": "UK利益率",
            "AG": "US送料上限",
            "AH": "UK内部価格"
        }

        print(f"\n--- 行 {target_row} ---")
        for col, label in columns.items():
            cell = f"{col}{target_row}"
            formula = ws[cell].value
            print(f"[{label} ({cell})]: {formula}")

    wb.close()
except Exception as e:
    print(f"Exception: {e}")
