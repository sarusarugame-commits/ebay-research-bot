import openpyxl

def extract_practice_formulas(file_path):
    sheet_name = "practice"
    
    # Excelファイルを読み込む（数式をそのまま取得するため、data_only=Falseを使用）
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except FileNotFoundError:
        print(f"エラー: '{file_path}' が見つかりません。ファイル名と保存場所を確認してください。")
        return None

    if sheet_name not in wb.sheetnames:
        print(f"エラー: '{sheet_name}' シートが見つかりません。")
        return None

    ws = wb[sheet_name]
    formulas = {}

    # シート内のすべてのセルをループして数式を探す
    for row in ws.iter_rows():
        for cell in row:
            # セルの値が文字列で、かつ "=" から始まる場合は数式とみなす
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[cell.coordinate] = cell.value

    return formulas

# 実行部分（ご自身のPCにあるExcelファイル名に合わせて変更してください）
file_name = "リサーチシート完全版 のコピー.xlsx" 

print(f"ファイル '{file_name}' の [{sheet_name}] シートを解析中...\n")
formulas_dict = extract_practice_formulas(file_name)

if formulas_dict is not None:
    if len(formulas_dict) > 0:
        print(f"--- [practice] シートの計算式一覧 (全 {len(formulas_dict)} 件) ---")
        # すべての数式を出力します
        for coord, formula in formulas_dict.items():
            print(f"セル {coord}: {formula}")
    else:
        print("このシートには数式が設定されたセルが見つかりませんでした。")