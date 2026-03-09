import openpyxl
import os

file_path = r"G:\マイドライブ\Python_code\eBayリサーチ部隊\リサーチシート完全版 のコピー.xlsx"
output_path = "debug_excel_analysis.txt"

try:
    wb_f = openpyxl.load_workbook(file_path, data_only=False)
    wb_v = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet_name = "3月" if "3月" in wb_f.sheetnames else "3月 のコピー"
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(f"Investigating sheet: {sheet_name}\n\n")

        target_row = 2
        cols = ["G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "X", "AG", "AH", "AI"]
        
        f.write(f"--- Row {target_row} Analysis ---\n")
        f.write(f"{'Col':<4} | {'Value':<12} | {'Formula'}\n")
        f.write("-" * 80 + "\n")
        for c in cols:
            cell = f"{c}{target_row}"
            val = ws_v[cell].value
            formula = ws_f[cell].value
            f.write(f"{c:<4} | {str(val):<12} | {formula}\n")

    wb_f.close()
    wb_v.close()
    print(f"Analysis saved to {output_path}")
except Exception as e:
    print(f"Error: {e}")
