import datetime
import unicodedata
import openpyxl
from openpyxl.styles import Font

EXCEL_PATH = r"G:\マイドライブ\Python_code\eBayリサーチ部隊\リサーチシート完全版 のコピー.xlsx"
SHEET_NAME = "3月 のコピー"

TARGET_MARGIN = 0.06

US_SHIPPING_TABLE = {
    500: 2060, 1000: 3221.4, 1500: 3433.3, 2000: 3606.2,
    3000: 4386.2, 4000: 4982.9, 5000: 6047.6, 6000: 6997.9,
    7000: 8045.7, 8000: 8513.7, 9000: 8983, 10000: 11264.5
}
UK_SHIPPING_TABLE = {
    500: 1571, 1000: 2908.1, 1500: 3383.9, 2000: 3802.5,
    3000: 4253.6, 4000: 4699.5, 5000: 5452.2, 6000: 6605.3,
    7000: 7127.9, 8000: 7650.5, 9000: 8174.4, 10000: 10067.2
}

def to_half_width(text):
    if text is None: return ""
    return unicodedata.normalize('NFKC', str(text))

def calculate_shipping_cost(weight_g, length_cm, width_cm, height_cm, table):
    """N列/Q列と同じ送料計算: max(実重量, 容積重量=縦×横×高さ/5) × 1.3"""
    try:
        vol_weight = float(length_cm) * float(width_cm) * float(height_cm) / 5.0
        billed_weight = max(float(weight_g), vol_weight) * 1.3
        for threshold in sorted(table.keys()):
            if billed_weight <= threshold:
                return table[threshold]
        return table[10000]
    except Exception:
        return table[500]

def simulate_k_us(h):
    """US利益率計算用: スプシのKをHで置き換え"""
    return h

def simulate_k_uk(j):
    """UK利益率計算用: スプシのKをJで置き換え"""
    return j

def calc_margin_us(k, l, exchange_rate, cost_jpy, shipping_jpy, is_high_tariff):
    """M列: O / ((K+L)*為替)"""
    tariff_rate = 0.4 if is_high_tariff else 0.2
    profit = (k * 0.9 + l) * exchange_rate * 0.8 \
             - k * 0.9 * tariff_rate * exchange_rate \
             - cost_jpy - shipping_jpy
    denom = (k + l) * exchange_rate
    return (profit / denom) if denom else 0.0, profit

def calc_margin_uk(k, exchange_rate, cost_jpy, shipping_jpy):
    """P列: R / (K*為替)  R=K*0.9*為替*0.8 - G - Q  ※関税なし・L列なし"""
    profit = k * 0.9 * exchange_rate * 0.8 - cost_jpy - shipping_jpy
    denom = k * exchange_rate
    return (profit / denom) if denom else 0.0, profit

def find_min_k_for_us(exchange_rate, cost_jpy, us_shipping_jpy, is_high_tariff, target=TARGET_MARGIN):
    """US利益率6%達成の最低K値を逆算（L=0仮定）"""
    tariff_rate = 0.4 if is_high_tariff else 0.2
    denom = exchange_rate * (0.9 * (0.8 - tariff_rate) - target)
    if denom <= 0:
        return None
    return (cost_jpy + us_shipping_jpy) / denom

def find_min_k_for_uk(exchange_rate, cost_jpy, uk_shipping_jpy, target=TARGET_MARGIN):
    """UK利益率6%達成の最低K値を逆算（関税なし・L=0）"""
    denom = exchange_rate * (0.72 - target)
    if denom <= 0:
        return None
    return (cost_jpy + uk_shipping_jpy) / denom

def write_to_sheet(item_data):
    print("\n[*] Excelシートへの書き込み処理を開始します...")

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[SHEET_NAME]
        wb_data = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws_data = wb_data[SHEET_NAME]
        exchange_rate = ws_data.cell(row=2, column=35).value
        if not isinstance(exchange_rate, (int, float)) or exchange_rate == 0:
            exchange_rate = 150.0
            print(f"    [!] 為替レート取得失敗。デフォルト {exchange_rate} 円/USD を使用。")
        else:
            print(f"    [*] 為替レート: {exchange_rate} 円/USD")
    except Exception as e:
        print(f"    [!] Excelファイルの読み込み失敗: {e}")
        return False

    def mm_to_cm(val):
        try: return float(val) / 10.0
        except: return 0.0

    def safe_float(val):
        try: return float(val)
        except: return 0.0

    length_cm = mm_to_cm(item_data.get('length', 0))
    width_cm  = mm_to_cm(item_data.get('width', 0))
    height_cm = mm_to_cm(item_data.get('height', 0))
    weight_g  = safe_float(item_data.get('weight', 0))
    cost      = float(item_data['domestic_price'])
    is_high_tariff = item_data.get('is_high_tariff', False)

    us_shipping_jpy = calculate_shipping_cost(weight_g, length_cm, width_cm, height_cm, US_SHIPPING_TABLE)
    uk_shipping_jpy = calculate_shipping_cost(weight_g, length_cm, width_cm, height_cm, UK_SHIPPING_TABLE)
    print(f"    [*] 国際送料試算: US=¥{us_shipping_jpy:,.0f} / UK=¥{uk_shipping_jpy:,.0f}")

    us_top3 = item_data.get('us_top3_prices', [])
    uk_top3 = item_data.get('uk_top3_prices', [])
    us_top3_shipping = item_data.get('us_top3_shipping', [])

    h_usd = min(us_top3) if us_top3 else None
    j_usd = min(uk_top3) if uk_top3 else None

    if h_usd is None and j_usd is not None:
        h_usd = j_usd
    if j_usd is None and h_usd is not None:
        j_usd = h_usd
    if h_usd is None and j_usd is None:
        print("    [SKIP] US/UKともに競合データなし。書き込みをスキップします。")
        return False

    # H/J それぞれ独立したKで利益率を仮計算
    # KはH/Jそれぞれの価格で計算（安い方がスプシのK列基準になる）
    k_us = simulate_k_us(h_usd)
    k_uk = simulate_k_uk(j_usd)
    m_us, profit_us = calc_margin_us(k_us, 0, exchange_rate, cost, us_shipping_jpy, is_high_tariff)
    m_uk, profit_uk = calc_margin_uk(k_uk, exchange_rate, cost, uk_shipping_jpy)
    print(f"    [試算] K_US=${k_us:.2f} K_UK=${k_uk:.2f} | US利益率: {m_us:.1%} | UK利益率: {m_uk:.1%}")

    # Top3の上限値（3件未満は上限なし）
    h_max = max(us_top3) if len(us_top3) >= 3 else float('inf')
    j_max = max(uk_top3) if len(uk_top3) >= 3 else float('inf')

    # 利益率未達の場合: H/J それぞれ独立して引き上げ
    if m_us < TARGET_MARGIN or m_uk < TARGET_MARGIN:
        k_min_us = find_min_k_for_us(exchange_rate, cost, us_shipping_jpy, is_high_tariff)
        k_min_uk = find_min_k_for_uk(exchange_rate, cost, uk_shipping_jpy)

        if k_min_us is None or k_min_uk is None:
            print(f"    [SKIP] 利益率{TARGET_MARGIN:.0%}の達成が構造的に不可能。スキップします。")
            return False

        h_min = round(k_min_us, 2)
        j_min = round(k_min_uk, 2)

        us_achievable = (h_min <= h_max)
        uk_achievable = (j_min <= j_max)

        if not us_achievable and not uk_achievable:
            print(f"    [SKIP] US(必要${h_min:.2f}>上限${h_max:.2f})・UK(必要${j_min:.2f}>上限${j_max:.2f})ともに6%達成不可。")
            return False

        # H/J をそれぞれ独立して引き上げ
        if us_achievable:
            h_usd = max(h_usd, h_min)
        if uk_achievable:
            j_usd = max(j_usd, j_min)

        k_us = simulate_k_us(h_usd)
        k_uk = simulate_k_uk(j_usd)
        m_us, profit_us = calc_margin_us(k_us, 0, exchange_rate, cost, us_shipping_jpy, is_high_tariff)
        m_uk, profit_uk = calc_margin_uk(k_uk, exchange_rate, cost, uk_shipping_jpy)
        print(f"    [ADJUST] H=${h_usd:.2f}(US最低${h_min:.2f}) J=${j_usd:.2f}(UK最低${j_min:.2f})")
        print(f"    [ADJUST] US利益率: {m_us:.1%} | UK利益率: {m_uk:.1%}")

    if m_us < TARGET_MARGIN and m_uk < TARGET_MARGIN:
        print(f"    [SKIP] US/UKともに利益率{TARGET_MARGIN:.0%}未満。スキップします。")
        return False

    print(f"    [確定] H=${h_usd:.2f} / J=${j_usd:.2f}")
    print(f"    [確定] US利益率: {m_us:.1%} (¥{profit_us:,.0f}) | UK利益率: {m_uk:.1%} (¥{profit_uk:,.0f})")

    # 書き込み先探索（B列が空の最初の行）
    target_row = None
    for row in range(2, ws.max_row + 100):
        if not ws.cell(row=row, column=2).value:
            target_row = row
            break
    if not target_row:
        target_row = ws.max_row + 1

    today_str = datetime.datetime.now().strftime("%m/%d").lstrip("0").replace("/0", "/")
    operator_info = f"{today_str}池田"

    def to_num(val):
        try: return float(val)
        except: return ""

    updates = {
        2:  item_data['product_name'],
        3:  round(mm_to_cm(item_data['length']), 1) if item_data.get('length') else "",
        4:  round(mm_to_cm(item_data['width']), 1)  if item_data.get('width')  else "",
        5:  round(mm_to_cm(item_data['height']), 1) if item_data.get('height') else "",
        6:  to_num(item_data.get('weight')),
        7:  item_data['domestic_price'],
        8:  h_usd,
        10: j_usd,
        22: item_data['source_url'],
        24: "高関税" if item_data.get('is_high_tariff') else "低関税",
        26: item_data['condition'],
        30: operator_info
    }

    font_10 = Font(size=10)
    for col_idx, val in updates.items():
        cell = ws.cell(row=target_row, column=col_idx)
        if isinstance(val, (int, float)):
            cell.value = val
        else:
            cell.value = to_half_width(val) if val != "" else val
        cell.font = font_10

    try:
        wb.save(EXCEL_PATH)
        print(f"    [SUCCESS] {target_row} 行目に書き込み完了。")
        return True
    except PermissionError:
        print(f"    [!] Excelファイルが開かれています。閉じてから再実行してください。")
        return False
